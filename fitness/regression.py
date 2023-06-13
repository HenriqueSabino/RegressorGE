from algorithm.parameters import params
from fitness.base_ff_classes.base_ff import base_ff
import requests
from sewar.full_ref import mse
from openpyxl import load_workbook
import numpy as np
import torch
import torch.nn.functional as F
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')


def extremeVal(reference, val):
    if (val > 1e2*reference) or (val < -1e2*reference):
        return 0
    else:
        return val


class regression(base_ff):

    # maximise = False
    # multi_objective = True

    def __init__(self):
        # Initialise base fitness function class.
        super().__init__()

        # self.num_obj = 2
        # dummyfit = base_ff()
        # dummyfit.maximise = False
        # self.fitness_functions = [dummyfit, dummyfit]
        # self.default_fitness = [float('nan'), float('nan')]

        self.split_type_index = {"arima": 2, "sarima": 1}
        self.models_values_column = {
            "arima": {
                "series": 1,
                "arima": 3,
                "mlp": 4,
                "svr": 5,
                "rbf": 6,
            },
            "sarima": {
                "series": 6,
                "sarima": 2,
                "mlp": 3,
                "svr": 4,
                "rbf": 5,
            },
        }

        dataset_name = params["DATASET_NAME"]
        print(f"DATASET_NAME: {dataset_name}")
        self.arima_dataset = self.load_data(
            params["DATASET_NAME"], params["ARIMA_DATASET_PATH"], "arima")

        self.sarima_dataset = self.load_data(
            params["DATASET_NAME"], params["SARIMA_DATASET_PATH"], "sarima")

    def get_metrics(self, phenotype):

        accuracy, accuracy_sd, f1_score, f1_score_sd = None, None, None, None

        r = requests.get(params['METRICS_URL'], params={
            'dataset': params['DATASET_PATH'],
            'phenotype': phenotype,
        })
        data = r.json()

        if len(data):
            data = data[0]
            accuracy = float(data['accuracy'])
            accuracy_sd = float(data['accuracy_sd'])
            f1_score = float(data['f1_score'])
            f1_score_sd = float(data['f1_score_sd'])

        return accuracy, accuracy_sd, f1_score, f1_score_sd

    def save_metrics(self, phenotype, accuracy, accuracy_sd, f1_score, f1_score_sd):
        data = {
            'dataset': params['DATASET_PATH'],
            'phenotype': phenotype,
            'accuracy': accuracy,
            'accuracy_sd': accuracy_sd,
            'f1_score': f1_score,
            'f1_score_sd': f1_score_sd,
        }
        requests.post(params['METRICS_URL'], json=data)

    def load_data(self, dataset_name, dataset_path, linear_type):
        wb = load_workbook(dataset_path)
        dataset_name_arr = [dataset_name]

        dataset = {}
        for name in dataset_name_arr:

            ws = wb[name]
            columns = [ws["A"], ws["B"], ws["C"],
                       ws["D"], ws["E"], ws["F"], ws["G"]]

            sheet = [[] for _ in range(len(columns[0]))]

            for c in columns:
                for index, item in enumerate(c):
                    sheet[index].append(item.value)

            sheet = np.array(sheet)
            dataset[name] = {"raw": sheet}

        for name, _ in dataset.items():
            sheet = dataset[name]["raw"]

            dataset[name]["series"] = {
                "treino": [], "teste": [], "validacao": []}
            dataset[name][linear_type] = {
                "treino": [], "teste": [], "validacao": []}
            dataset[name]["mlp"] = {"treino": [], "teste": [], "validacao": []}
            dataset[name]["svr"] = {"treino": [], "teste": [], "validacao": []}
            dataset[name]["rbf"] = {"treino": [], "teste": [], "validacao": []}

            for row in sheet[1:]:
                split_type = row[self.split_type_index[linear_type]]
                if linear_type == "arima" and row[4] == None:
                    continue
                key = None
                if split_type.lower() == "teste":
                    key = "teste"
                elif split_type.lower() == "validacao" or split_type.lower() == "validação":
                    key = "validacao"
                elif split_type.lower() == "treinamento":
                    key = "treino"
                else:
                    continue

                for model_key, val in self.models_values_column[linear_type].items():
                    dataset[name][model_key][key].append(float(row[val]))

            for type_data in ["teste", "treino", "validacao"]:
                dataset[name]["series"][type_data] = np.array(
                    dataset[name]["series"][type_data])
                dataset[name][linear_type][type_data] = np.array(
                    dataset[name][linear_type][type_data])
                dataset[name]["mlp"][type_data] = np.array(
                    dataset[name]["mlp"][type_data])
                dataset[name]["svr"][type_data] = np.array(
                    dataset[name]["svr"][type_data])
                dataset[name]["rbf"][type_data] = np.array(
                    dataset[name]["rbf"][type_data])

        fixExtreme = np.vectorize(extremeVal)

        for name in dataset.keys():
            reference = np.max(dataset[name]["series"]["treino"])

            for type_data in ["teste", "treino", "validacao"]:
                dataset[name]["mlp"][type_data] = fixExtreme(
                    reference, dataset[name]["mlp"][type_data])
                dataset[name]["svr"][type_data] = fixExtreme(
                    reference, dataset[name]["svr"][type_data])
                dataset[name]["rbf"][type_data] = fixExtreme(
                    reference, dataset[name]["rbf"][type_data])
                dataset[name][linear_type][type_data] = fixExtreme(
                    reference, dataset[name][linear_type][type_data])

        return dataset[dataset_name]

    def build_model(self, phenotype):
        model = {"linear": {}, "nlinear": {}}

        linear, nlinear = phenotype.split(';')

        weight_tuple, linear_model = linear.split(':')

        weight_values = weight_tuple.removeprefix('(').removesuffix(')')
        model["linear"][linear_model] = [
            float(x) for x in weight_values.split(' ')]

        nlinear_parts = nlinear.split(',')

        for i in range(len(nlinear_parts)):
            weight_tuple, nlinear_model = nlinear_parts[i].split(':')
            weight_values = weight_tuple.removeprefix('(').removesuffix(')')

            if nlinear_model in model["nlinear"]:
                model["nlinear"][nlinear_model] += [
                    float(x) for x in weight_values.split(' ')]
            else:
                model["nlinear"][nlinear_model] = [
                    float(x) for x in weight_values.split(' ')]

        return model

    def predict_mse(self, model, split):

        if "arima" in model["linear"].keys():
            dataset = self.arima_dataset
            linear_type = "arima"
        else:
            dataset = self.sarima_dataset
            linear_type = "sarima"

        predict = dataset["series"][split]

        window_size = params["WINDOW_SIZE"]
        X_train = self.apply_window(
            window_size, split, linear_type, dataset, model_names=model["nlinear"].keys())

        kernel = []
        for i in range(window_size):
            for _, val in model["linear"].items():
                kernel.append(val[i])

            for _, val in model["nlinear"].items():
                kernel.append(val[i])

        return mse(self.create_prediction(X_train, np.reshape(kernel, (1, -1))), predict)

    def evaluate(self, ind, **kwargs):
        model = self.build_model(ind.phenotype)
        train_mse = self.predict_mse(model, "treino")
        # validation_mse = self.predict_mse(model, "validacao")

        # if (train_mse - validation_mse) < 0:
        #     return train_mse * 10
        # if (train_mse - validation_mse) > 0:
        #     return 0.1 * train_mse

        return train_mse

    def apply_window(self, window_size, data_type, linear_type, dataset, model_names=None):

        size_pred = len(dataset["series"][data_type])
        new_data = []

        for w in range(window_size):
            linear = np.concatenate(
                [np.zeros(w), dataset[linear_type][data_type]])
            linear = linear[:size_pred]
            row = np.array(linear)
            for model_name in model_names:
                nonlinear = np.concatenate(
                    [np.zeros(w), dataset[model_name][data_type]])
                nonlinear = nonlinear[:size_pred]
                row = np.column_stack([row, nonlinear])

            if len(new_data) == 0:
                new_data = row
            else:
                new_data = np.column_stack([new_data, row])

        return np.array(new_data)

    def create_prediction(self, X, kernel):

        X = torch.tensor(np.expand_dims(X, axis=(0, 1))).float()
        kernel = torch.tensor(np.expand_dims(kernel, axis=(0, 1))).float()

        X.to(device)
        kernel.to(device)

        result = F.conv2d(X, kernel)
        return result.numpy().squeeze().astype("f")

    @staticmethod
    def value(fitness_vector, objective_index):
        """
        This is a static method required by NSGA-II for sorting populations
        based on a given fitness function, or for returning a given index of a
        population based on a given fitness function.

        :param fitness_vector: A vector/list of fitnesses.
        :param objective_index: The index of the desired fitness.
        :return: The fitness at the objective index of the fitness vector.
        """

        if not isinstance(fitness_vector, list):
            return float("inf")

        return fitness_vector[objective_index]
